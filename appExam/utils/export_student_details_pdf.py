from io import BytesIO

import qrcode
from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import Image
from reportlab.platypus import PageBreak
from reportlab.platypus import SimpleDocTemplate
from reportlab.platypus import Spacer
from reportlab.platypus import Table
from reportlab.platypus import TableStyle

from appExam.models import ExamSession
from appExam.models import SeatAssignment
from appExam.models import StudentExamEnrollment


def generate_qr_code(symbol_number: str) -> BytesIO:
    """Generate QR code for a symbol number and return as BytesIO."""
    qr = qrcode.QRCode(
        version=1,
        box_size=10,
        border=1,
    )
    qr.add_data(symbol_number)
    qr.make(fit=True)
    img = qr.make_image(fill="black", back_color="white")

    img_buffer = BytesIO()
    img.save(img_buffer, format="PNG")
    img_buffer.seek(0)
    return img_buffer


def download_exam_pdf_view(request, session_id):
    session = ExamSession.objects.get(pk=session_id)
    enrollments = StudentExamEnrollment.objects.filter(session=session).select_related(
        "candidate",
    )

    buffer = BytesIO()
    page_width, page_height = A4
    left_margin = right_margin = 40
    usable_width = page_width - left_margin - right_margin

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=20,  # shift content higher
        bottomMargin=20,
    )

    elements = []

    for enrollment in enrollments:
        candidate = enrollment.candidate
        seat = (
            SeatAssignment.objects.filter(enrollment=enrollment)
            .select_related("hall")
            .first()
        )

        if seat and seat.seat_number:
            hall_name = seat.hall.name if seat.hall else "No Hall"
            if isinstance(seat.seat_number, int):
                formatted_seat_number = f"{hall_name}-C{seat.seat_number:03d}"
            else:
                prefix = "".join([c for c in seat.seat_number if c.isalpha()])
                number_part = "".join([c for c in seat.seat_number if c.isdigit()])
                formatted_seat_number = (
                    f"{hall_name} - {prefix}{int(number_part):03d}"
                    if number_part
                    else f"{hall_name} - {seat.seat_number}"
                )
        else:
            formatted_seat_number = "Not Assigned"

        password = getattr(candidate, "generated_password", "N/A")

        # Prepare table data
        data = [
            ["Symbol No", "Password", "Seat Number"],
            [candidate.symbol_number, password, formatted_seat_number],
        ]
        col_widths = [usable_width * 0.33, usable_width * 0.33, usable_width * 0.34]

        table = Table(data, colWidths=col_widths)
        table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Courier-Bold"),
                    ("FONTNAME", (0, 1), (-1, -1), "Courier"),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ],
            ),
        )

        # Generate QR code image
        qr_buffer = generate_qr_code(candidate.symbol_number)
        qr_image = Image(qr_buffer, width=100, height=100)  # Adjust size as needed

        # Layout: Table at top, then QR code below with small spacing
        elements.append(table)
        elements.append(Spacer(1, 12))  # small gap
        elements.append(qr_image)
        elements.append(PageBreak())

    # Remove trailing page break
    if elements and isinstance(elements[-1], PageBreak):
        elements.pop()

    doc.build(elements)
    buffer.seek(0)

    return FileResponse(
        buffer,
        as_attachment=True,
        filename=f"Exam_Session_{session.exam.program.name}_{session.base_start}_Enrollments.pdf",
    )


def download_exam_excel_view(request, session_id):
    session = ExamSession.objects.get(pk=session_id)
    enrollments = StudentExamEnrollment.objects.filter(session=session).select_related(
        "candidate",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Exam Enrollments"

    # Header row
    ws.append(["Symbol No", "Password", "Seat Number"])

    for enrollment in enrollments:
        candidate = enrollment.candidate
        seat = (
            SeatAssignment.objects.filter(enrollment=enrollment)
            .select_related("hall")
            .first()
        )

        if seat and seat.seat_number:
            hall_name = seat.hall.name if seat.hall else "No Hall"
            if isinstance(seat.seat_number, int):
                formatted_seat_number = f"{hall_name}-C{seat.seat_number:03d}"
            else:
                prefix = "".join([c for c in seat.seat_number if c.isalpha()])
                number_part = "".join([c for c in seat.seat_number if c.isdigit()])
                formatted_seat_number = (
                    f"{hall_name} - {prefix}{int(number_part):03d}"
                    if number_part
                    else f"{hall_name} - {seat.seat_number}"
                )
        else:
            formatted_seat_number = "Not Assigned"

        # Get password without replacing any characters
        password = getattr(candidate, "generated_password", "N/A")

        # Append row
        ws.append(
            [
                candidate.symbol_number,
                password,
                formatted_seat_number,
            ],
        )

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save workbook to bytes buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Exam_Session_{session.exam.program.name}_{session.base_start}_Enrollments.xlsx"

    return FileResponse(
        output,
        as_attachment=True,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )